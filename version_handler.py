import os
import re
import pandas as pd
import logging
import shutil
from openpyxl import load_workbook

from config import (VERSION,
                    HISTORY_DB_PATH, EFFICIENCY_REPORT_PATH,
                    BACKUP_HISTORY_DB_PATH, BACKUP_EFFICIENCY_REPORT_PATH,
                    EMPLOYEE_DB_PATH, STOCK_DB_PATH)
from utils import file_access_handler
from data_processor import calculate_financial_metrics


def regenerate_history_from_dbs():
    """从 EmployeeDB 和 StockDB 重新生成 HISTORY_DB"""
    print("🔄 开始重建 HISTORY_DB...")
    logging.info("开始从 EmployeeDB 和 StockDB 重建 HISTORY_DB")

    try:
        # 读取所有 sheet
        df_emp = pd.read_excel(EMPLOYEE_DB_PATH, sheet_name=None)
        df_stock = pd.read_excel(STOCK_DB_PATH, sheet_name=None)

        old_df_report = pd.read_excel(
            BACKUP_EFFICIENCY_REPORT_PATH, sheet_name=None)
        old_df_report = next(iter(old_df_report.values()))
        first_col_name = old_df_report.columns[0]
        old_df_report = old_df_report.set_index(first_col_name)
        old_df_report.columns = [pd.to_datetime(col).strftime('%Y-%m-%d') if col != '职位' else col
                                 for col in old_df_report.columns]

        print(f"从 EmployeeDB 读取到 {len(df_emp)} 个日期数据")
        print(f"从 StockDB 读取到 {len(df_stock)} 个日期数据")

        # 获取所有日期并排序
        all_dates = sorted(set(df_emp.keys()) & set(df_stock.keys()))

        history_list = []

        for date_str in all_dates:
            current_df_emp = df_emp[date_str]
            current_df_stock = df_stock[date_str]

            # 计算财务指标
            gross_income = int(old_df_report.loc['日收入', date_str])
            advertising_budget = int(
                old_df_report.loc['广告费 (M)', date_str] * 1000000)

            financial_metrics = calculate_financial_metrics(
                current_df_emp, current_df_stock, advertising_budget, gross_income)

            # 生成员工数据
            if 'eff_total' not in current_df_emp.columns:
                current_df_emp['eff_total'] = 0

            df_sorted = current_df_emp.sort_values(
                ['position', 'eff_total'], ascending=[True, False])
            df_sorted['pos_rank'] = df_sorted.groupby(
                'position').cumcount() + 1
            df_sorted['position'] = df_sorted['position'].astype(
                str) + df_sorted['pos_rank'].astype(str)

            daily_data = df_sorted[['position', 'eff_total']].copy()
            daily_data.rename(
                columns={'eff_total': 'Efficiency_Sum'}, inplace=True)
            daily_data['日期'] = date_str

            # 添加财务指标（按你希望的顺序）
            metrics = pd.DataFrame([
                {"position": "合计效率", "Efficiency_Sum": current_df_emp['eff_total'].sum(
                ), "日期": date_str},
                {"position": "环境 (%)", "Efficiency_Sum": old_df_report.loc['环境 (%)', date_str],
                 "日期": date_str},
                {"position": "日收入", "Efficiency_Sum": gross_income, "日期": date_str},
                {"position": "工资支出",
                    "Efficiency_Sum": financial_metrics['total_employee_salary'], "日期": date_str},
                {"position": "广告费", "Efficiency_Sum": advertising_budget, "日期": date_str},
                {"position": "销售成本",
                    "Efficiency_Sum": financial_metrics['total_stock_cost'], "日期": date_str},
                {"position": "净利润",
                    "Efficiency_Sum": financial_metrics['net_profit'], "日期": date_str}
            ])

            daily_data = pd.concat([daily_data, metrics], ignore_index=True)
            history_list.append(daily_data)

        # 合并所有日期数据
        full_history = pd.concat(history_list, ignore_index=True)

        # 保存
        full_history.to_excel(HISTORY_DB_PATH, index=False)
        print(f"✅ HISTORY_DB 重建完成，共 {len(all_dates)} 天数据")
        logging.info(f"✅ HISTORY_DB 重建完成，共 {len(all_dates)} 天数据")
        return True
    except Exception as e:
        logging.error(f"重建 HISTORY_DB 失败: {e}")
        print(f"❌ 重建失败: {e}")
        return False


@file_access_handler
def check_industry_version_upgrade(sheet_name, industry_db_path):
    """检查 IndustryDB 中指定 Sheet 是否为旧 API 截断数据（正好100条），若是则删除旧 Sheet"""
    if not os.path.exists(industry_db_path):
        return

    try:
        wb = load_workbook(industry_db_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return

        ws = wb[sheet_name]
        # max_row - 1 排除表头
        if ws.max_row - 1 == 100:
            logging.warning(
                f"IndustryDB Sheet {sheet_name} 仅有 100 条数据（旧 API v1 限制），将删除并重新写入")
            print(
                f"⚠️ IndustryDB Sheet {sheet_name} 仅有 100 条数据（旧 API v1 限制），将更新为完整数据")
            del wb[sheet_name]
            wb.save(industry_db_path)
            wb.close()
            print(f"✅ 已删除旧 Sheet: {sheet_name}")
            logging.info(f"已删除旧 IndustryDB Sheet: {sheet_name}")
        else:
            wb.close()
    except Exception as e:
        logging.error(f"检查 IndustryDB 版本升级失败: {e}")


@file_access_handler
def check_and_upgrade_report():
    """检查 Efficiency_Report 是否为新版本，如果不是则升级"""
    if not os.path.exists(EFFICIENCY_REPORT_PATH):
        logging.info("Efficiency_Report.xlsx 不存在，跳过版本检查")
        return False

    try:
        xls = pd.ExcelFile(EFFICIENCY_REPORT_PATH)
        sheet_names = xls.sheet_names
        old_sheet_name = sheet_names[0] if sheet_names else None
        print(f"🔍 检测到 Efficiency_Report.xlsx，当前版本: {old_sheet_name or '未知'}")

        # 已是最新版本
        if old_sheet_name == f"v{VERSION}":
            logging.info(f"✅ Efficiency_Report 已是最新版本 (Sheet: {f'v{VERSION}'})")
            xls.close()
            return True
        xls.close()

        # v1.5 之后格式：Sheet 名以 vX.X 开头，只需重命名
        if old_sheet_name and re.match(r'^v\d+\.\d+$', old_sheet_name):
            logging.warning(f"⚠️ 检测到旧版本 {old_sheet_name}，只需重命名 Sheet...")
            print(f"⚠️ 检测到旧版本 {old_sheet_name}，正在升级 Sheet 名称...")
            wb = load_workbook(EFFICIENCY_REPORT_PATH)
            wb[old_sheet_name].title = f"v{VERSION}"
            wb.save(EFFICIENCY_REPORT_PATH)
            wb.close()
            logging.info(f"✅ 报表 Sheet 已从 {old_sheet_name} 升级到 v{VERSION}")
            print(f"✅ 报表 Sheet 已从 {old_sheet_name} 升级到 v{VERSION}")
            return True

        # v1.4 及之前：完整备份并重建
        logging.warning("⚠️ 检测到旧版本 Efficiency_Report（非 vX.X 格式），正在进行完整升级...")

        shutil.copy2(HISTORY_DB_PATH, BACKUP_HISTORY_DB_PATH)
        shutil.copy2(EFFICIENCY_REPORT_PATH, BACKUP_EFFICIENCY_REPORT_PATH)
        print(f"✅ 旧文件已备份：")
        print(f"   • {BACKUP_HISTORY_DB_PATH}")
        print(f"   • {BACKUP_EFFICIENCY_REPORT_PATH}")

        # 删除旧文件
        os.remove(HISTORY_DB_PATH)
        os.remove(EFFICIENCY_REPORT_PATH)

        # === 重新生成 HISTORY_DB ===
        print("🔄 正在根据最新数据重新生成 HISTORY_DB...")
        regenerate_history_from_dbs()

        logging.info("✅ 版本升级完成")
        return True

    except Exception as e:
        logging.error(f"版本检查失败: {e}")
        return False


