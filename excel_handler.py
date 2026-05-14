import os
import pandas as pd
import logging
from openpyxl import load_workbook

from utils import file_access_handler


@file_access_handler
def save_to_excel(file_path, df, sheet_name, replace=False, header=True):
    if df.empty:
        logging.warning(f"数据为空，跳过保存: {file_path} - {sheet_name}")
        return False

    try:
        use_header = header
        if len(df.columns) == 0 or (df.columns[0] is None):
            use_header = False

        if replace or not os.path.exists(file_path):
            df.to_excel(file_path, sheet_name=sheet_name,
                        index=False, header=use_header)
            print(
                f"✅ {file_path} 中 {sheet_name} 已{'创建' if not os.path.exists(file_path) else '更新'}")
            return True

        wb = load_workbook(file_path)
        sheet_exists = sheet_name in wb.sheetnames
        wb.close()

        if sheet_exists and not replace:
            print(f"⚠️ 跳过：{file_path} 中已存在 Sheet: {sheet_name}")
            return False

        mode = 'a' if os.path.exists(file_path) else 'w'
        with pd.ExcelWriter(file_path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name,
                        index=False, header=use_header)

        print(f"✅ {file_path} 中 {sheet_name} 已保存")
        return True

    except Exception as e:
        logging.error(f"保存失败: {e}")
        return False


@file_access_handler
def generate_horizontal_report(history_file, output_file, target_date):
    """根据历史记录生成横向对比报表"""
    logging.info(f"开始生成横向效率对比报表 - 日期: {target_date}")

    if not os.path.exists(history_file):
        logging.warning(f"历史文件不存在，跳过报表生成: {history_file}")
        return

    target_date_str = str(target_date)

    # 检查输出文件是否已包含今日日期
    if os.path.exists(output_file):
        existing_report = pd.read_excel(output_file)
        existing_cols = [str(col).split(' ')[0]
                         for col in existing_report.columns]
        if target_date_str in existing_cols:
            print(f"ℹ️ 报表记录 [{target_date_str}] 已存在，跳过生成。")
            logging.info(f"报表 [{target_date_str}] 已存在，跳过生成")
            return

    df = pd.read_excel(history_file)
    df['日期'] = pd.to_datetime(df['日期']).dt.date

    pivot_df = df.pivot(index='position', columns='日期',
                        values='Efficiency_Sum').reset_index()
    pivot_df.rename(columns={'position': '职位'}, inplace=True)

    date_cols = sorted([col for col in pivot_df.columns if col != '职位'])
    pivot_df = pivot_df[['职位'] + date_cols]

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        pivot_df.to_excel(writer, sheet_name='Efficiency_Report', index=False)
        ws = writer.sheets['Efficiency_Report']
        for col in ws.columns:
            max_len = max([len(str(cell.value)) for cell in col])
            ws.column_dimensions[col[0].column_letter].width = min(
                max_len + 2, 25)

    logging.info(f"横向对比报表生成成功: {output_file}")
